# first install this package by command below:
# $ pip install openpyxl

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

my_work_book = xl.load_workbook('transactions.xlsx')

sheet = my_work_book['Sheet1']

print(sheet)

cell = sheet['a1']
print(cell)
print(cell.value)

# read from .xlsx
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        print(sheet.cell(row, column).value, end='\t')
    print()


# write in .xlsx
sheet.cell(10, 10).value = "10"
my_work_book.save("transactions_append.xlsx")



